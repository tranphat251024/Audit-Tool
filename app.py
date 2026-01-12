import streamlit as st
import google.generativeai as genai
import pdfplumber
from docx import Document
from PIL import Image
import openpyxl
import json
import os
import fitz  # PyMuPDF
import re
import time

# --- 1. CẤU HÌNH & BẢO MẬT ---
st.set_page_config(page_title="Auto-Audit V16: Secured", page_icon="🔐", layout="wide")

# MẬT KHẨU ĐỂ VÀO TOOL (Bạn hãy sửa lại theo ý thích)
ADMIN_PASSWORD = "admin" 

RULE_FILE = "saved_rules.json"

# --- 2. HỆ THỐNG ĐĂNG NHẬP ĐƠN GIẢN ---
def check_password():
    """Trả về True nếu chưa đăng nhập, False nếu đã đăng nhập thành công"""
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False
    
    if not st.session_state['logged_in']:
        st.markdown("## 🔐 ĐĂNG NHẬP HỆ THỐNG BẢO MẬT")
        pwd = st.text_input("Nhập mật khẩu quản trị:", type="password")
        if st.button("Đăng nhập"):
            if pwd == ADMIN_PASSWORD:
                st.session_state['logged_in'] = True
                st.success("Đăng nhập thành công!")
                time.sleep(0.5)
                st.rerun()
            else:
                st.error("Sai mật khẩu!")
        return True # Vẫn đang ở màn hình khóa
    return False # Đã mở khóa

# NẾU CHƯA ĐĂNG NHẬP THÌ DỪNG LẠI, KHÔNG CHẠY CODE BÊN DƯỚI
if check_password():
    st.stop()

# =========================================================
# TỪ ĐÂY TRỞ XUỐNG LÀ CODE LOGIC V15 (ĐÃ ĐƯỢC BẢO VỆ)
# =========================================================

# --- HÀM QUẢN LÝ ---
def load_rules():
    if os.path.exists(RULE_FILE):
        try:
            with open(RULE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return []
    return []

def save_new_rule(rule):
    rules = load_rules()
    if rule and rule not in rules:
        rules.append(rule)
        with open(RULE_FILE, "w", encoding="utf-8") as f:
            json.dump(rules, f, ensure_ascii=False, indent=4)
    return rules

def clear_rules():
    if os.path.exists(RULE_FILE):
        os.remove(RULE_FILE)
    return []

# --- BỘ ĐỌC FILE ---
def process_single_file(uploaded_file):
    if not uploaded_file: return None
    uploaded_file.seek(0)
    try:
        if uploaded_file.type in ['image/jpeg', 'image/png', 'image/jpg']:
            img = Image.open(uploaded_file)
            if img.width > 1500: img.thumbnail((1500, 1500))
            return img
        elif uploaded_file.name.endswith('.pdf'):
            text = ""
            with pdfplumber.open(uploaded_file) as pdf:
                for page in pdf.pages:
                    extract = page.extract_text()
                    if extract: text += extract + "\n"
            return text if text.strip() else f"[Scan PDF: {uploaded_file.name}]"
        elif uploaded_file.name.endswith('.docx'):
            doc = Document(uploaded_file)
            return "\n".join([para.text for para in doc.paragraphs])
        elif uploaded_file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                text += f"\n--- SHEET EXCEL: {sheet} ---\n"
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip(): text += row_text + "\n"
            return text
        return None
    except Exception as e: return f"Error: {e}"

def process_multiple_files(file_list):
    combined_text = ""
    collected_images = []
    if not file_list: return "", []
    for file in file_list:
        result = process_single_file(file)
        if isinstance(result, str):
            combined_text += f"\n=== NỘI DUNG TỪ FILE: {file.name} ===\n{result}\n"
        elif isinstance(result, Image.Image):
            collected_images.append(result)
    return combined_text, collected_images

def highlight_errors_on_pdf(pdf_file, error_keywords):
    if not pdf_file or not error_keywords: return []
    pdf_file.seek(0)
    doc = fitz.open(stream=pdf_file.getvalue(), filetype="pdf")
    highlighted_images = []
    for page_num, page in enumerate(doc):
        found = False
        for kw in error_keywords:
            clean_kw = kw.strip().replace('"', '').replace("'", "")
            if len(clean_kw) < 2: continue
            quads = page.search_for(clean_kw)
            if quads:
                found = True
                for quad in quads:
                    page.draw_rect(quad, color=(1, 0, 0), width=2, fill=(1, 0, 0), fill_opacity=0.2)
        if found:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_data = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            highlighted_images.append((page_num + 1, img_data))
    return highlighted_images

def get_working_model():
    try:
        for m in genai.list_models():
            if '1.5-flash' in m.name and 'generateContent' in m.supported_generation_methods: return m.name
        models = genai.list_models()
        for m in models:
            if 'generateContent' in m.supported_generation_methods: return m.name
    except: return None
    return 'models/gemini-1.5-flash'

# --- ENGINE V6 + TEMPLATE ---
def analyze_secured(tpl_txt, tpl_imgs, s_txt, s_imgs, t_txt, t_imgs, rules, api_key):
    if not api_key: yield "⚠️ Thiếu API Key!"; return

    genai.configure(api_key=api_key)
    yield "🔐 Đang chạy trong môi trường bảo mật (Temp=0)...\n"
    
    model_name = get_working_model()
    if not model_name: yield "❌ Lỗi kết nối Model."; return
    
    model = genai.GenerativeModel(
        model_name, 
        generation_config=genai.types.GenerationConfig(temperature=0.0, top_p=1.0)
    )
    
    rules_text = "\n".join([f"- {r}" for r in rules])
    
    template_instruction = ""
    if tpl_txt or tpl_imgs:
        template_instruction = f"""
        PHẦN 0: KIỂM TRA TUÂN THỦ FILE MẪU
        - Kiểm tra Báo cáo có tuân thủ cấu trúc/số liệu của File Mẫu không.
        - Text mẫu: {tpl_txt[:5000]}...
        """
    else:
        template_instruction = "(Không có File Mẫu)."

    instruction = f"""
    Bạn là Auditor nghiêm ngặt.
    LUẬT MIỄN TRỪ: {rules_text}
    
    {template_instruction}
    
    THỰC HIỆN 4 BƯỚC KIỂM TRA SỐ LIỆU (V6 STANDARD):
    
    BƯỚC 1: ĐỊNH DANH (CCCD, Tên, Năm sinh).
    BƯỚC 2: PHÁP LÝ (Tờ, Thửa, Diện tích, Địa chỉ).
    BƯỚC 3: HIỆN TRẠNG (Ảnh vs Mô tả).
    BƯỚC 4: GIÁ & LOGIC.
    
    TRẢ KẾT QUẢ:
    1. Liệt kê lỗi sai.
    2. Bọc từ sai trong [[[...]]] (VD: [[[100m2]]]).
    3. Đúng ghi "✅ Khớp".
    
    TARGET: {t_txt}
    SOURCE: {s_txt}
    """
    
    content = [instruction]
    if tpl_imgs: content.append("\nTEMPLATE IMAGES:"); content.extend(tpl_imgs)
    if s_imgs: content.append("\nSOURCE IMAGES:"); content.extend(s_imgs)
    if t_imgs: content.append("\nTARGET IMAGES:"); content.extend(t_imgs)

    try:
        response = model.generate_content(content, stream=True)
        for chunk in response:
            if chunk.text: yield chunk.text
    except Exception as e: yield f"❌ Lỗi: {str(e)}"

# --- GIAO DIỆN CHÍNH ---
st.title("🔐 Auto-Audit V16: Secured System")

# Nút Đăng xuất
if st.sidebar.button("🔒 Đăng xuất / Khóa màn hình"):
    st.session_state['logged_in'] = False
    st.rerun()

st.markdown("---")

if 'rules' not in st.session_state: st.session_state['rules'] = load_rules()

with st.sidebar:
    st.header("⚙️ Cấu Hình")
    # Gợi ý: Bạn có thể nhập sẵn key vào đây để đỡ phải gõ
    # api_key = "AIzaSy....." 
    api_key = st.text_input("API Key:", type="password")
    
    st.markdown("---")
    st.subheader("📂 File Mẫu (Templates)")
    f_tpls = st.file_uploader("Upload Mẫu (Excel/Word/Ảnh)", type=['pdf','docx','xlsx','png','jpg'], accept_multiple_files=True, key="tpl")
    if f_tpls:
        with st.spinner("Học mẫu..."):
            tpl_txt, tpl_imgs = process_multiple_files(f_tpls)
            st.session_state['tpl_txt'] = tpl_txt
            st.session_state['tpl_imgs'] = tpl_imgs
        st.success("✅ Đã học mẫu!")
    else: st.session_state['tpl_txt'] = ""; st.session_state['tpl_imgs'] = []

    st.markdown("---")
    with st.expander(f"🧠 Bộ nhớ luật ({len(st.session_state['rules'])})"):
        for r in st.session_state['rules']: st.write(f"- {r}")
    if st.button("Xóa bộ nhớ"): st.session_state['rules'] = clear_rules(); st.rerun()

c1, c2 = st.columns(2)
with c1:
    st.info("📂 1. DỮ LIỆU NGUỒN")
    f_src = st.file_uploader("Upload Nguồn", type=['pdf','docx','xlsx','png','jpg'], accept_multiple_files=True, key="u1")
    if f_src:
        with st.spinner("Đọc nguồn..."): s_txt, s_imgs = process_multiple_files(f_src)
        st.success(f"Nhận {len(s_imgs)} ảnh.")
    else: s_txt, s_imgs = "", []

with c2:
    st.info("📝 2. BÁO CÁO")
    f_tgt = st.file_uploader("Upload Báo cáo (PDF)", type=['pdf'], accept_multiple_files=True, key="u2")
    if f_tgt:
        with st.spinner("Đọc báo cáo..."): t_txt, t_imgs = process_multiple_files(f_tgt)
        st.success(f"Nhận {len(f_tgt)} file.")
    else: t_txt, t_imgs = "", []

st.markdown("---")

if st.button("🚀 BẮT ĐẦU KIỂM TRA", type="primary", use_container_width=True):
    if (not s_txt and not s_imgs) or (not t_txt):
        st.warning("⚠️ Thiếu dữ liệu!")
    else:
        st.markdown("### 📊 Kết quả")
        with st.expander("📄 Xem chi tiết", expanded=True):
            res_box = st.empty()
            full_report = ""
            for chunk in analyze_secured(
                st.session_state.get('tpl_txt', ''), 
                st.session_state.get('tpl_imgs', []), 
                s_txt, s_imgs, t_txt, t_imgs, 
                st.session_state['rules'], api_key
            ):
                full_report += chunk
                res_box.markdown(full_report + "▌")
            res_box.markdown(full_report)
            st.session_state['last_report'] = full_report
        
        error_keywords = re.findall(r"\[\[\[(.*?)\]\]\]", full_report)
        st.markdown("---")
        col_L, col_R = st.columns(2)
        with col_L:
            st.subheader("⬅️ Dữ liệu Gốc")
            if s_imgs:
                tabs = st.tabs([f"Ảnh {i+1}" for i in range(len(s_imgs))])
                for i, t in enumerate(tabs):
                    with t: st.image(s_imgs[i], use_container_width=True)
            else: 
                with st.expander("Xem text nguồn"): st.write(s_txt)
        with col_R:
            st.subheader("➡️ Báo Cáo (Khoanh lỗi)")
            if f_tgt and error_keywords:
                pdf_files = [f for f in f_tgt if f.name.endswith('.pdf')]
                found = False
                with st.container(height=600):
                    for pdf in pdf_files:
                        highlighted = highlight_errors_on_pdf(pdf, error_keywords)
                        if highlighted:
                            found = True
                            st.caption(f"File: {pdf.name}")
                            for p, img in highlighted: st.image(img, caption=f"Trang {p}", use_container_width=True)
                if not found: st.warning("Không tìm thấy vị trí lỗi.")
            elif not error_keywords: st.success("✅ Sạch lỗi.")

if st.session_state.get('last_report'):
    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1: st.download_button("📥 Tải Báo Cáo", st.session_state['last_report'], "Audit_Result.txt", use_container_width=True)
    with c2: 
        with st.form("teach"):
            r = st.text_input("Dạy luật mới:"); submit = st.form_submit_button("Lưu")
            if submit: st.session_state['rules'] = save_new_rule(r); st.rerun()